"""
Export router: Excel export of client interaction logs.

Architecture Decision:
- openpyxl for Excel generation (no subprocess, no shell injection risk)
- Stream the file as a response without writing to disk
- Date range filtering prevents exporting unbounded data
- Max 10,000 rows per export to prevent memory exhaustion
"""

import io
from datetime import datetime, timezone
from typing import Optional

import structlog
from fastapi import APIRouter, Depends, Query, Request, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.exceptions import AppException
from app.core.security import TokenData, get_current_user
from app.repositories.repositories import AuditRepository, QueryRepository

router = APIRouter()
logger = structlog.get_logger(__name__)

MAX_EXPORT_ROWS = 10_000


@router.get(
    "/queries",
    summary="Export query logs as Excel file",
)
async def export_queries(
    request: Request,
    start_date: Optional[datetime] = Query(None, description="ISO8601 start date"),
    end_date: Optional[datetime] = Query(None, description="ISO8601 end date"),
    current_user: TokenData = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Export all RAG queries for the client as an Excel file.
    Optionally filter by date range.
    Maximum 10,000 rows per export.
    """
    query_repo = QueryRepository(db)
    audit_repo = AuditRepository(db)

    queries = await query_repo.get_all_for_export(
        current_user.client_id,
        start_date=start_date,
        end_date=end_date,
    )

    if len(queries) > MAX_EXPORT_ROWS:
        raise AppException(
            f"Export exceeds maximum of {MAX_EXPORT_ROWS} rows. Please narrow your date range.",
            status_code=400,
        )

    await audit_repo.log(
        action="EXPORT",
        client_id=current_user.client_id,
        user_id=current_user.user_id,
        ip_address=request.headers.get("X-Forwarded-For"),
        request_id=getattr(request.state, "request_id", None),
        extra={"row_count": len(queries)},
    )

    # Build Excel workbook
    wb = _build_workbook(queries, str(current_user.client_id))

    # Stream to client as bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"query_logs_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

    logger.info(
        "Export generated",
        client_id=str(current_user.client_id),
        row_count=len(queries),
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_workbook(queries: list, client_id: str) -> Workbook:
    """Build formatted Excel workbook from query data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Query Logs"

    # Header style
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Query ID",
        "Timestamp",
        "Question",
        "Answer",
        "Status",
        "Tokens Used",
        "Latency (ms)",
        "Sources",
    ]

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, q in enumerate(queries, start=2):
        source_filenames = ""
        if q.source_documents:
            filenames = [s.get("filename", "") for s in q.source_documents if s.get("filename")]
            source_filenames = "; ".join(set(filenames))

        ws.cell(row=row_idx, column=1, value=str(q.id))
        ws.cell(row=row_idx, column=2, value=q.created_at.isoformat() if q.created_at else "")
        ws.cell(row=row_idx, column=3, value=q.question[:1000] if q.question else "")
        ws.cell(row=row_idx, column=4, value=q.answer[:2000] if q.answer else "")
        ws.cell(row=row_idx, column=5, value=q.status)
        ws.cell(row=row_idx, column=6, value=q.tokens_used)
        ws.cell(row=row_idx, column=7, value=q.latency_ms)
        ws.cell(row=row_idx, column=8, value=source_filenames)

    # Set column widths
    column_widths = [38, 22, 60, 80, 12, 14, 14, 40]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary["A1"] = "Export Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = "Client ID:"
    ws_summary["B3"] = client_id
    ws_summary["A4"] = "Total Queries:"
    ws_summary["B4"] = len(queries)
    ws_summary["A5"] = "Export Date:"
    ws_summary["B5"] = datetime.now(timezone.utc).isoformat()

    success_count = sum(1 for q in queries if q.status == "success")
    ws_summary["A6"] = "Successful:"
    ws_summary["B6"] = success_count
    ws_summary["A7"] = "Failed:"
    ws_summary["B7"] = len(queries) - success_count

    total_tokens = sum(q.tokens_used or 0 for q in queries)
    ws_summary["A8"] = "Total Tokens Used:"
    ws_summary["B8"] = total_tokens

    return wb
